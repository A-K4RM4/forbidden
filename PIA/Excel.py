def funcionExcel():
   import os, re, requests, json, time
   from bs4 import BeautifulSoup as bs
   from datetime import datetime
    #El grupo (0)es todo el mensaje de Reynolds/grupo(1) es la fecha/grupos(2,3,4)corresponden a d/m/a /grupo(5) es el lugar,(6 y 7) son la ciudad y estado
   patnDR=re.compile(r' naciÃ³ el ((\d{2}) de (\w+) de (\d{4})) en ((\w+) (\w+))')
    #Grupo(1) es la fecha de Platzman/Gpos(2,3,4) son d/m/a 
   patnDJ=re.compile(r' nacido el ((\d{2}) de (\w+) de (\d{4}))')
    #Grupo(1) aqui corresponde al lugar de nacimiento de Platzman
   patlDJ=re.compile(r'naciÃ³ en ((\w+), (\w+))')
    #Grupo(1) fecha nacimiento de Benjamin
   patnBAM=re.compile(r'naciÃ³ el ((\d{2}) de (\w+) de (\d{4}))')
    #Grupo(1) lugar de nacimineto de Ben
   patlBAM=re.compile(r' en ((\w+) en (\w+))')
    #Ojo aqui, el grupo 1 es el lugar y el grupo 4 es la fecha, Wayne es el unico que tiene un orden de busqueda de su info distinto
   patnDW=re.compile(r'naciÃ³ en ((\w+) (\w+)) el ((\d{2}) de (\w+) de (\d{4}))')
        
   r = requests.get('https://www.imaginedragonsmusic.com/')
   soup = bs(r.content,"html.parser") #Clase beautifulSoup
    #Creamos un archivo html para referencia
   fo = open ("C:/Paginas/Pagina1.txt","w", encoding="utf-8")
   fo.write (str(soup.prettify()))
   fo.close()

    #Cerramos el archivo html
   r = requests.get('https://imagine-dragons.fandom.com/es/wiki/Dan_Reynolds')
   soup = bs(r.content,"html.parser") #Clase beautifulSoup
    #Creamos un archivo html para referencia
   fo = open ("C:/Paginas/Integrante1.txt","w", encoding="utf-8")
   fo.write (str(soup.prettify()))
   fo.close()
    #Cerramos el archivo html
    #Leemos el txt para sacar informacion
   with open ("C:/Paginas/Integrante1.txt",'r') as file:
      for line in file:
         if " naciÃ³ el" in line:
               patron=patnDR.search(line)
               fechaDR=patron.group(1)
               lugarDR=patron.group(5)
               break

   r = requests.get('https://imagine-dragons.fandom.com/es/wiki/Daniel_Platzman')
   soup = bs(r.content,"html.parser") #Clase beautifulSoup
    #Creamos un archivo html para referencia
   fo = open ("C:/Paginas/Integrante2.txt","w", encoding="utf-8")
   fo.write (str(soup.prettify()))
   fo.close()
    #Cerramos el archivo html
    #Leemos el txt para sacar informacion
   with open ("C:/Paginas/Integrante2.txt",'r') as file:
      for line in file:
         if "nacido el" in line:
               patron=patnDJ.search(line)
               fechaDJ=patron.group(1)
               break
   with open ("C:/Paginas/Integrante2.txt",'r') as file:
      for line in file:
         if "naciÃ³ en" in line:
               patl=patlDJ.search(line)
               lugarDJ=patl.group(1)
               break

   r = requests.get('https://imagine-dragons.fandom.com/es/wiki/Ben_McKee')
   soup = bs(r.content,"html.parser") #Clase beautifulSoup
    #Creamos un archivo html para referencia
   fo = open ("C:/Paginas/Integrante3.txt","w", encoding="utf-8")
   fo.write (str(soup.prettify()))
   fo.close()
    #Cerramos el archivo html
    #Leemos el txt para sacar informacion
   with open ("C:/Paginas/Integrante3.txt",'r') as file:
      for line in file:
         if " naciÃ³ el" in line:
               patron=patnBAM.search(line)
               fechaBAM=patron.group(1)
               break
   with open ("C:/Paginas/Integrante3.txt",'r') as file:
      for line in file:
         if "Ben viviÃ³" in line:
               patlb=patlBAM.search(line)
               lugarBAM=patlb.group(1)
               break
    
   r = requests.get('https://imagine-dragons.fandom.com/es/wiki/Wayne_Sermon')
   soup = bs(r.content,"html.parser") #Clase beautifulSoup
    #Creamos un archivo html para referencia
   fo = open ("C:/Paginas/Integrante4.txt","w", encoding="utf-8")
   fo.write (str(soup.prettify()))
   fo.close()
    #Cerramos el archivo html
    #Leemos el txt para sacar informacion
   with open ("C:/Paginas/Integrante4.txt",'r') as file:
      for line in file:
         if " naciÃ³ en" in line:
               patron=patnDW.search(line)
               lugarDW=patron.group(1)
               fechaDW=patron.group(4)
               break

   appid=input("Ingresa tu APPID: ")
    #Ámsterdam - Concierto 1 
   lat = "52.3740311"
   lon = "4.8896899"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
    #Aqui sacamos la temperatura del día
   for elem in weatherData["current"]:
      if elem == "temp":
         T1=weatherData["current"][elem]
    #Aqui obtenemos la probabilidad de lluvia
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL1= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL1= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL1= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL1= "Probabilidad de lluvia: 85-100%"
    #Aqui obtenemos la salida y entrada del sol
   dt = 1591900805
   SS1=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES1=time.strftime("%H:%M:%S",time.localtime(dt))

    #Paris
   lat = "48.8534088"
   lon = "2.3487999"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T2=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL2= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL2= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL2= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL2= "Probabilidad de lluvia: 85-100%"         
   dt = 1591900805
   SS2=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES2=time.strftime("%H:%M:%S",time.localtime(dt))

    #Nueva York
   lat = "40.66"
   lon = "-73.93"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T3=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL3= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL3= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL3= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL3= "Probabilidad de lluvia: 85-100%"         
   dt = 1591996452
   SS3=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES3=time.strftime("%H:%M:%S",time.localtime(dt))
          
    #Buenos Aires
   lat = "-34.6131516"
   lon = "-58.3772316"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T4=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL4= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL4= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL4= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL4= "Probabilidad de lluvia: 85-100%"         
   dt = 1592082852
   SS4=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES4=time.strftime("%H:%M:%S",time.localtime(dt))

    #Madrid
   lat = "40.4165"
   lon = "-3.7025"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T5=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL5= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL5= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL5= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL5= "Probabilidad de lluvia: 85-100%"         
   dt = 1592082852
   SS5=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES5=time.strftime("%H:%M:%S",time.localtime(dt))
   
    #Monterrey
   lat = "25.6714"
   lon = "-100.309"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T6=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL6= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL6= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL6= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL6= "Probabilidad de lluvia: 85-100%"         
   dt = 1592169252
   SS6=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES6=time.strftime("%H:%M:%S",time.localtime(dt))

    #Londres
   lat = "51.5072"
   lon = "-0.1275"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T7=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL7= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL7= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL7= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL7= "Probabilidad de lluvia: 85-100%"         
   dt = 1592169252
   SS7=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES7=time.strftime("%H:%M:%S",time.localtime(dt))

    #Tokio
   lat = "35.6894"
   lon = "139.692"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T8=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL8= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL8= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL8= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL8= "Probabilidad de lluvia: 85-100%"         
   dt = 1592255652
   SS8=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES8=time.strftime("%H:%M:%S",time.localtime(dt))

    #Bruselas
   lat = "50.85044"
   lon = "4.348780"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T9=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL9= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL9= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL9= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL9= "Probabilidad de lluvia: 85-100%"         
   dt = 1592342052
   SS9=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES9=time.strftime("%H:%M:%S",time.localtime(dt))

    #São Paulo
   lat = "-23.5475"
   lon = "-46.6361"
   page= requests.get("https://api.openweathermap.org/data/2.5/onecall?lat="+lat+"&lon="+lon+"&exclude=minutely,hourly,daily&units=metric&appid="+appid)
   weatherData=json.loads(page.content)
   for elem in weatherData["current"]:
      if elem == "temp":
         T10=weatherData["current"][elem]
   for key in weatherData["current"]["weather"][0]:
      if key == "id":
         if (weatherData["current"]["weather"][0][key]) == 801:
               PLL10= "Probabilidad de lluvia: 11-25%"
         elif (weatherData["current"]["weather"][0][key]) == 802:
               PLL10= "Probabilidad de lluvia: 25-50%"
         elif (weatherData["current"]["weather"][0][key]) == 803:
               PLL10= "Probabilidad de lluvia: 51-84%"
         elif (weatherData["current"]["weather"][0][key]) == 804:
               PLL10= "Probabilidad de lluvia: 85-100%"         
   dt = 1592428452
   SS10=datetime.utcfromtimestamp(dt).strftime("%H:%M:%S")
   ES10=time.strftime("%H:%M:%S",time.localtime(dt))

   try:
      from openpyxl import load_workbook, Workbook
   except ImportError:
       os.system('pip install openpyxl')
       print('Instalando Openpyxl. El programa cerrará, vuelva a ejecutar el código')
       exit()

   try:
       libro = load_workbook("Imagine Dragons.xlsx")
   except FileNotFoundError:
       print("El archivo no existe\nCreando archivo\nVuelva a ejecutar el programa!")
       LibroID = Workbook()
       LibroID.save("Imagine Dragons.xlsx")
       exit()
       
   LibroID = Workbook()
   Hoja = LibroID.active
   Hoja.title = "Datos" 
   LibroID.save("Imagine Dragons.xlsx")

   if len(LibroID.sheetnames)>1:
       LibroID.active = 1
       Hoja = LibroID.active
   else:
       Hoja = LibroID.active

   Hoja["A1"] = "Nombre"
   Hoja["B1"] = "Fecha de Nacimiento"
   Hoja["C1"] = "Lugar de Nacimiento"
   #Llenamos la primer columna
   Hoja["A2"] = "Dan Reynolds"
   Hoja["A3"] = "Daniel Platzman"
   Hoja["A4"] = "Ben McKee"
   Hoja["A5"] = "Daniel Wayne Sermon"
   #Llenamos la segunda columna con los datos obtenidos con RegExP
   Hoja["B2"] = fechaDR
   Hoja["B3"] = fechaDJ
   Hoja["B4"] = fechaBAM
   Hoja["B5"] = fechaDW
   #Llenamos la tercera columna con los datos obtenidos con RegExP
   Hoja["C2"] = lugarDR
   Hoja["C3"] = lugarDJ
   Hoja["C4"] = lugarBAM
   Hoja["C5"] = lugarDW
   
   #Aqui asignamos los nombres de las columnas de la tabla del concierto
   Hoja["E1"] = "Fecha del Concierto"
   Hoja["F1"] = "Lugar de Concierto"
   Hoja["G1"] = "Clima"
   Hoja["H1"] = "Probabilidad de lluvia"
   Hoja["I1"] = "Salida del Sol"
   Hoja["J1"] = "Entrada del Sol"
    #Llenamos la primer columna
   Hoja["E2"] = "11 de Junio"
   Hoja["E3"] = "11 de Junio"
   Hoja["E4"] = "12 de Junio"
   Hoja["E5"] = "13 de Junio"
   Hoja["E6"] = "13 de Junio"
   Hoja["E7"] = "14 de Junio"
   Hoja["E8"] = "14 de Junio"
   Hoja["E9"] = "15 de Junio"
   Hoja["E10"] = "16 de Junio"
   Hoja["E11"] = "17 de Junio"
   #Llenamos la segunda columna con los datos obtenidos con RegExP
   Hoja["F2"] = "Ámsterdam"
   Hoja["F3"] = "París"
   Hoja["F4"] = "Nueva York"
   Hoja["F5"] = "Buenos Aires"
   Hoja["F6"] = "Madrid"
   Hoja["F7"] = "Monterrey"
   Hoja["F8"] = "Londres"
   Hoja["F9"] = "Tokio"
   Hoja["F10"] = "Bruselas"
   Hoja["F11"] = "São Paulo"
   #Llenamos la tercera columna con los datos obtenidos con RegExP
   Hoja["G2"] = T1
   Hoja["G3"] = T2
   Hoja["G4"] = T3
   Hoja["G5"] = T4
   Hoja["G6"] = T5
   Hoja["G7"] = T6
   Hoja["G8"] = T7
   Hoja["G9"] = T8
   Hoja["G10"] = T9
   Hoja["G11"] = T10
   #Llenamos la cuarta columna con los datos obtenidos con RegExp
   Hoja["H2"] = PLL1
   Hoja["H3"] = "No hay información"
   Hoja["H4"] = "No hay información"
   Hoja["H5"] = PLL4
   Hoja["H6"] = PLL5
   Hoja["H7"] = PLL6
   Hoja["H8"] = PLL7
   Hoja["H9"] = PLL8
   Hoja["H10"] = "No hay información"
   Hoja["H11"] = PLL10
   #Llenamos la quinta columna con los datos obtenidos con RegExP
   Hoja["I2"] = SS1
   Hoja["I3"] = SS2
   Hoja["I4"] = SS3
   Hoja["I5"] = SS4
   Hoja["I6"] = SS5
   Hoja["I7"] = SS6
   Hoja["I8"] = SS7
   Hoja["I9"] = SS8
   Hoja["I10"] = SS9
   Hoja["I11"] = SS10
   #Llenamos la sexta columna con los datos obtenidos con RegExP
   Hoja["J2"] = ES1
   Hoja["J3"] = ES2
   Hoja["J4"] = ES3
   Hoja["J5"] = ES4
   Hoja["J6"] = ES5
   Hoja["J7"] = ES6
   Hoja["J8"] = ES7
   Hoja["J9"] = ES8
   Hoja["J10"] = ES9
   Hoja["J11"] = ES10


   LibroID.save("Imagine Dragons.xlsx")
   LibroID.close()
